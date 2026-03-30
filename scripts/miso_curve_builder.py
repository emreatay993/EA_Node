from __future__ import annotations

import json
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Sequence

try:
    from PyQt6.QtCore import QPointF, QSignalBlocker, Qt, pyqtSignal
    from PyQt6.QtGui import QColor, QKeySequence, QPainter, QPen
    from PyQt6.QtWidgets import (
        QApplication,
        QAbstractItemView,
        QComboBox,
        QFileDialog,
        QGridLayout,
        QGroupBox,
        QHBoxLayout,
        QHeaderView,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPushButton,
        QSplitter,
        QTableWidget,
        QTableWidgetItem,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )

    QT_API = "PyQt6"
except ImportError:  # pragma: no cover - runtime fallback for user machines
    from PyQt5.QtCore import QPointF, QSignalBlocker, Qt, pyqtSignal
    from PyQt5.QtGui import QColor, QKeySequence, QPainter, QPen
    from PyQt5.QtWidgets import (
        QApplication,
        QAbstractItemView,
        QComboBox,
        QFileDialog,
        QGridLayout,
        QGroupBox,
        QHBoxLayout,
        QHeaderView,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPushButton,
        QSplitter,
        QTableWidget,
        QTableWidgetItem,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )

    QT_API = "PyQt5"

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - surfaced through UI message
    Workbook = None


ALIGNMENT = getattr(Qt, "AlignmentFlag", Qt)
ITEM_FLAG = getattr(Qt, "ItemFlag", Qt)
KEY_ENUM = getattr(Qt, "Key", Qt)
RESIZE_MODE = getattr(QHeaderView, "ResizeMode", QHeaderView)
SELECTION_BEHAVIOR = getattr(QAbstractItemView, "SelectionBehavior", QAbstractItemView)
SELECTION_MODE = getattr(QAbstractItemView, "SelectionMode", QAbstractItemView)
EDIT_TRIGGER = getattr(QAbstractItemView, "EditTrigger", QAbstractItemView)
RENDER_HINT = getattr(QPainter, "RenderHint", QPainter)
PASTE_SHORTCUT = getattr(QKeySequence, "StandardKey", QKeySequence).Paste
COPY_SHORTCUT = getattr(QKeySequence, "StandardKey", QKeySequence).Copy
HEADER_STRETCH = RESIZE_MODE.Stretch
SELECT_ITEMS = SELECTION_BEHAVIOR.SelectItems
SELECT_ROWS = SELECTION_BEHAVIOR.SelectRows
SINGLE_SELECTION = SELECTION_MODE.SingleSelection
EXTENDED_SELECTION = SELECTION_MODE.ExtendedSelection
NO_EDIT_TRIGGERS = EDIT_TRIGGER.NoEditTriggers
KEY_DELETE = KEY_ENUM.Key_Delete
KEY_BACKSPACE = KEY_ENUM.Key_Backspace

MIN_TEMPERATURE_ROWS = 8
MIN_CURVE_ROWS = 18
TEMPERATURE_KEY_DIGITS = 9
PROJECT_FILE_VERSION = 1
DEFAULT_PROOF_OFFSET_MODE = "proof_0p2"


@dataclass
class RawCurveRow:
    stress_text: str = ""
    strain_text: str = ""


@dataclass
class TemperatureCurveData:
    temperature_c: float
    raw_rows: list[RawCurveRow] = field(default_factory=list)
    manual_modulus_text: str = ""


@dataclass
class ConversionSettings:
    curve_type: str
    modulus_mode: str
    manual_modulus_mpa: float | None
    proof_offset_mode: str
    proof_offset_plastic_strain: float


@dataclass
class ParsedClipboard:
    rows: list[list[str]]
    accepted_count: int
    skipped_count: int
    header_skipped: bool


@dataclass
class ParsedCurveInput:
    rows: list[RawCurveRow]
    accepted_count: int
    skipped_count: int
    header_skipped: bool


@dataclass
class TrueCurvePoint:
    strain_true: float
    stress_true: float
    input_strain_percent: float
    input_stress_mpa: float


@dataclass
class MisoPoint:
    plastic_strain: float
    stress_mpa: float


@dataclass(frozen=True)
class ElasticFitResult:
    modulus_mpa: float
    intercept_mpa: float
    r_squared: float
    rmse_mpa: float
    point_count: int
    strain_min: float
    strain_max: float
    stress_min_mpa: float
    stress_max_mpa: float
    max_secant_deviation: float
    points: tuple[TrueCurvePoint, ...]

    @property
    def strain_span(self) -> float:
        return self.strain_max - self.strain_min

    @property
    def intercept_ratio(self) -> float:
        return abs(self.intercept_mpa) / max(abs(self.stress_max_mpa), 1.0)


@dataclass
class ComputationResult:
    miso_rows: list[MisoPoint]
    displayed_curve: list[tuple[float, float]]
    modulus_mpa: float
    modulus_origin: str
    parse_summary: str
    elastic_fit: ElasticFitResult | None = None


class CurveValidationError(ValueError):
    pass


class ProjectFileError(ValueError):
    pass


def parse_localized_float(text: str) -> float:
    token = text.strip().replace("\u00a0", "").replace(" ", "")
    if not token:
        raise ValueError("Blank token.")
    if token.count(",") and token.count("."):
        if token.rfind(",") > token.rfind("."):
            token = token.replace(".", "").replace(",", ".")
        else:
            token = token.replace(",", "")
    elif token.count(",") == 1 and token.count(".") == 0:
        token = token.replace(",", ".")
    else:
        token = token.replace(",", "")
    return float(token)


def format_number(value: float, digits: int = 12) -> str:
    return f"{value:.{digits}g}"


def tooltip_html(title: str, *lines: str) -> str:
    body = "<br>".join(lines)
    return f"<b>{title}</b><br>{body}"


def proof_offset_value(mode: str) -> float:
    if mode == "proof_0p02":
        return 0.0002
    return 0.002


def proof_offset_label(mode: str) -> str:
    if mode == "proof_0p02":
        return "0.02%"
    return "0.2%"


def sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]", "_", name).strip()
    cleaned = cleaned or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 1
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def split_clipboard_line(line: str) -> list[str]:
    stripped = line.strip()
    if not stripped:
        return []
    if "\t" in stripped:
        return [cell.strip() for cell in stripped.split("\t")]
    return re.split(r"\s+", stripped)


def parse_numeric_clipboard(text: str, expected_columns: int) -> ParsedClipboard:
    rows: list[list[str]] = []
    skipped_count = 0
    header_skipped = False
    first_nonblank_seen = False
    for line in text.splitlines():
        tokens = split_clipboard_line(line)
        if not tokens:
            continue
        selected = tokens[:expected_columns]
        try:
            if len(selected) < expected_columns:
                raise ValueError("Too few columns.")
            numeric_values = [parse_localized_float(token) for token in selected]
        except ValueError:
            if not first_nonblank_seen:
                header_skipped = True
            else:
                skipped_count += 1
            first_nonblank_seen = True
            continue
        rows.append([format_number(value) for value in numeric_values])
        first_nonblank_seen = True
    return ParsedClipboard(
        rows=rows,
        accepted_count=len(rows),
        skipped_count=skipped_count,
        header_skipped=header_skipped,
    )


def parse_curve_rows(rows: Sequence[RawCurveRow]) -> ParsedCurveInput:
    accepted: list[RawCurveRow] = []
    skipped_count = 0
    header_skipped = False
    first_nonblank_seen = False
    for row in rows:
        stress_text = row.stress_text.strip()
        strain_text = row.strain_text.strip()
        if not stress_text and not strain_text:
            continue
        try:
            stress_value = parse_localized_float(stress_text)
            strain_value = parse_localized_float(strain_text)
        except ValueError:
            if not first_nonblank_seen:
                header_skipped = True
            else:
                skipped_count += 1
            first_nonblank_seen = True
            continue
        accepted.append(
            RawCurveRow(
                stress_text=format_number(stress_value),
                strain_text=format_number(strain_value),
            )
        )
        first_nonblank_seen = True
    return ParsedCurveInput(
        rows=accepted,
        accepted_count=len(accepted),
        skipped_count=skipped_count,
        header_skipped=header_skipped,
    )


def parse_temperature_table_rows(rows: Sequence[str]) -> ParsedClipboard:
    accepted: list[list[str]] = []
    skipped_count = 0
    header_skipped = False
    first_nonblank_seen = False
    for value in rows:
        text = value.strip()
        if not text:
            continue
        try:
            temperature = parse_localized_float(text)
        except ValueError:
            if not first_nonblank_seen:
                header_skipped = True
            else:
                skipped_count += 1
            first_nonblank_seen = True
            continue
        accepted.append([format_number(temperature)])
        first_nonblank_seen = True
    return ParsedClipboard(
        rows=accepted,
        accepted_count=len(accepted),
        skipped_count=skipped_count,
        header_skipped=header_skipped,
    )


def temperature_key(value: float) -> str:
    return f"{round(value, TEMPERATURE_KEY_DIGITS):.{TEMPERATURE_KEY_DIGITS}f}"


def format_temperature_label(value: float) -> str:
    return f"{format_number(value, 9)} C"


def linear_regression_slope(points: Sequence[TrueCurvePoint]) -> float:
    if len(points) < 2:
        raise CurveValidationError("At least two points are required to fit elastic modulus.")
    x_values = [point.strain_true for point in points]
    y_values = [point.stress_true for point in points]
    count = float(len(points))
    sum_x = sum(x_values)
    sum_y = sum(y_values)
    sum_xx = sum(value * value for value in x_values)
    sum_xy = sum(x * y for x, y in zip(x_values, y_values))
    denominator = count * sum_xx - sum_x * sum_x
    if abs(denominator) < 1e-18:
        first, second = points[0], points[1]
        delta_x = second.strain_true - first.strain_true
        if delta_x <= 0.0:
            raise CurveValidationError("Elastic modulus could not be inferred from duplicate strain values.")
        return (second.stress_true - first.stress_true) / delta_x
    return (count * sum_xy - sum_x * sum_y) / denominator


def linear_regression_fit(points: Sequence[TrueCurvePoint]) -> tuple[float, float, float, float]:
    slope = linear_regression_slope(points)
    x_values = [point.strain_true for point in points]
    y_values = [point.stress_true for point in points]
    count = float(len(points))
    intercept = (sum(y_values) - slope * sum(x_values)) / count
    residuals = [y - (slope * x + intercept) for x, y in zip(x_values, y_values)]
    ss_res = sum(residual * residual for residual in residuals)
    mean_y = sum(y_values) / count
    ss_tot = sum((y - mean_y) * (y - mean_y) for y in y_values)
    if ss_tot <= 1e-18:
        r_squared = 1.0
    else:
        r_squared = max(0.0, 1.0 - ss_res / ss_tot)
    rmse = math.sqrt(ss_res / count)
    return slope, intercept, r_squared, rmse


def convert_to_true_curve(rows: Sequence[RawCurveRow], curve_type: str) -> list[TrueCurvePoint]:
    points: list[TrueCurvePoint] = []
    for row in rows:
        stress = parse_localized_float(row.stress_text)
        strain_percent = parse_localized_float(row.strain_text)
        if stress < 0.0:
            raise CurveValidationError("Negative stress values are not supported for MISO export.")
        if curve_type == "engineering":
            engineering_strain = strain_percent / 100.0
            if engineering_strain <= -1.0:
                raise CurveValidationError("Engineering strain must stay above -100%.")
            stress_true = stress * (1.0 + engineering_strain)
            strain_true = math.log1p(engineering_strain)
        else:
            stress_true = stress
            strain_true = strain_percent / 100.0
        points.append(
            TrueCurvePoint(
                strain_true=strain_true,
                stress_true=stress_true,
                input_strain_percent=strain_percent,
                input_stress_mpa=stress,
            )
        )
    if len(points) < 2:
        raise CurveValidationError("At least two valid stress-strain rows are required.")
    return sorted(points, key=lambda point: (point.strain_true, point.stress_true))


def build_elastic_fit(points: Sequence[TrueCurvePoint]) -> ElasticFitResult:
    if len(points) < 2:
        raise CurveValidationError("Elastic modulus inference requires at least two points.")
    slope, intercept, r_squared, rmse = linear_regression_fit(points)
    if slope <= 0.0:
        raise CurveValidationError("Inferred elastic modulus must be positive.")
    secant_slopes: list[float] = []
    for point_a, point_b in zip(points, points[1:]):
        delta_strain = point_b.strain_true - point_a.strain_true
        if delta_strain <= 0.0:
            raise CurveValidationError("Elastic modulus inference requires increasing strain values.")
        secant_slope = (point_b.stress_true - point_a.stress_true) / delta_strain
        if secant_slope <= 0.0:
            raise CurveValidationError("Elastic modulus inference requires increasing stress values.")
        secant_slopes.append(secant_slope)
    max_secant_deviation = 0.0
    if secant_slopes:
        max_secant_deviation = max(abs(secant_slope - slope) / slope for secant_slope in secant_slopes)
    stresses = [point.stress_true for point in points]
    strains = [point.strain_true for point in points]
    return ElasticFitResult(
        modulus_mpa=slope,
        intercept_mpa=intercept,
        r_squared=r_squared,
        rmse_mpa=rmse,
        point_count=len(points),
        strain_min=min(strains),
        strain_max=max(strains),
        stress_min_mpa=min(stresses),
        stress_max_mpa=max(stresses),
        max_secant_deviation=max_secant_deviation,
        points=tuple(points),
    )


def fit_summary_text(fit: ElasticFitResult) -> str:
    return (
        f"{fit.point_count} pts, "
        f"{format_number(fit.strain_min * 100.0, 5)}% to {format_number(fit.strain_max * 100.0, 5)}% true strain, "
        f"R^2 = {format_number(fit.r_squared, 6)}, "
        f"intercept = {format_number(fit.intercept_mpa, 6)} MPa"
    )


def _fit_preference_key(fit: ElasticFitResult) -> tuple[float, ...]:
    return (
        float(fit.point_count),
        fit.strain_span,
        fit.stress_max_mpa,
        fit.r_squared,
        -fit.intercept_ratio,
        -fit.max_secant_deviation,
    )


def infer_elastic_modulus(points: Sequence[TrueCurvePoint]) -> ElasticFitResult:
    distinct_points: list[TrueCurvePoint] = []
    for point in points:
        if distinct_points and math.isclose(
            point.strain_true,
            distinct_points[-1].strain_true,
            rel_tol=1e-12,
            abs_tol=1e-12,
        ):
            continue
        distinct_points.append(point)
    if len(distinct_points) < 2:
        raise CurveValidationError("Elastic modulus inference requires two distinct strain values.")
    candidate_points = [point for point in distinct_points if point.strain_true <= 0.01]
    if len(candidate_points) < 2:
        candidate_points = distinct_points[: min(len(distinct_points), 12)]
    else:
        candidate_points = candidate_points[:12]
    if len(candidate_points) == 2:
        return build_elastic_fit(candidate_points)

    candidate_fits: list[ElasticFitResult] = []
    for start_index in range(len(candidate_points) - 1):
        for end_index in range(start_index + 1, len(candidate_points)):
            window = candidate_points[start_index : end_index + 1]
            try:
                fit = build_elastic_fit(window)
            except CurveValidationError:
                continue
            if fit.modulus_mpa <= 0.0:
                continue
            candidate_fits.append(fit)
    if not candidate_fits:
        raise CurveValidationError("Elastic modulus inference requires a monotonic elastic segment.")

    threshold_steps = (
        (4, 0.9995, 0.08, 0.015),
        (3, 0.9980, 0.12, 0.030),
        (3, 0.9950, 0.20, 0.050),
    )
    for min_points, min_r_squared, max_secant_deviation, max_intercept_ratio in threshold_steps:
        qualified = [
            fit
            for fit in candidate_fits
            if fit.point_count >= min_points
            and fit.r_squared >= min_r_squared
            and fit.max_secant_deviation <= max_secant_deviation
            and fit.intercept_ratio <= max_intercept_ratio
        ]
        if qualified:
            return max(qualified, key=_fit_preference_key)

    fallback = max(
        candidate_fits,
        key=lambda fit: (
            fit.r_squared,
            float(fit.point_count),
            fit.strain_span,
            -fit.intercept_ratio,
            -fit.max_secant_deviation,
        ),
    )
    if (
        fallback.point_count < 3
        or fallback.r_squared < 0.99
        or fallback.max_secant_deviation > 0.25
        or fallback.intercept_ratio > 0.08
    ):
        raise CurveValidationError(
            "Elastic modulus inference could not find a stable elastic window. Use Manual E or fit a cleaner elastic selection."
        )
    return fallback


def interpolate_stress_at_plastic_strain(candidate_a: MisoPoint, candidate_b: MisoPoint, target: float) -> float:
    delta_strain = candidate_b.plastic_strain - candidate_a.plastic_strain
    if abs(delta_strain) < 1e-18:
        raise CurveValidationError("Plastic strain proof point could not be interpolated.")
    factor = (target - candidate_a.plastic_strain) / delta_strain
    return candidate_a.stress_mpa + factor * (candidate_b.stress_mpa - candidate_a.stress_mpa)


def build_miso_rows(
    points: Sequence[TrueCurvePoint],
    modulus_mpa: float,
    proof_offset_plastic_strain: float,
) -> list[MisoPoint]:
    if modulus_mpa <= 0.0:
        raise CurveValidationError("Elastic modulus must be positive.")
    candidates = [
        MisoPoint(
            plastic_strain=point.strain_true - point.stress_true / modulus_mpa,
            stress_mpa=point.stress_true,
        )
        for point in points
    ]
    tolerance = 1e-6
    target_plastic_strain = proof_offset_plastic_strain
    positive_index = next(
        (
            index
            for index, candidate in enumerate(candidates)
            if candidate.plastic_strain > target_plastic_strain + tolerance
        ),
        None,
    )
    if positive_index is None:
        raise CurveValidationError(
            "No proof-stress crossing was found. Check the curve type, proof setting, or elastic modulus."
        )
    if positive_index == 0:
        raise CurveValidationError(
            "Plastic strain is already above the selected proof offset at the first row. Check the curve type or elastic modulus."
        )

    leading_candidates = candidates[:positive_index]
    last_nonpositive = leading_candidates[-1]
    if abs(last_nonpositive.plastic_strain - target_plastic_strain) <= tolerance:
        zero_row = MisoPoint(plastic_strain=0.0, stress_mpa=last_nonpositive.stress_mpa)
    else:
        previous_negative = next(
            (
                candidate
                for candidate in reversed(leading_candidates)
                if candidate.plastic_strain < target_plastic_strain - tolerance
            ),
            None,
        )
        if previous_negative is None:
            raise CurveValidationError(
                "No proof-stress crossing was found. Check the curve type, proof setting, or elastic modulus."
            )
        zero_row = MisoPoint(
            plastic_strain=0.0,
            stress_mpa=interpolate_stress_at_plastic_strain(
                previous_negative,
                candidates[positive_index],
                target_plastic_strain,
            ),
        )

    miso_rows = [zero_row]
    last_plastic_strain = zero_row.plastic_strain
    last_stress = zero_row.stress_mpa
    for candidate in candidates[positive_index:]:
        adjusted_plastic_strain = candidate.plastic_strain - target_plastic_strain
        if adjusted_plastic_strain <= last_plastic_strain + tolerance:
            continue
        if candidate.stress_mpa < last_stress - 1e-9:
            continue
        miso_rows.append(MisoPoint(plastic_strain=adjusted_plastic_strain, stress_mpa=candidate.stress_mpa))
        last_plastic_strain = adjusted_plastic_strain
        last_stress = candidate.stress_mpa
    return miso_rows


def compute_miso_from_rows(rows: Sequence[RawCurveRow], settings: ConversionSettings) -> ComputationResult:
    parsed_rows = parse_curve_rows(rows)
    if parsed_rows.accepted_count < 2:
        raise CurveValidationError("Paste at least two valid stress-strain rows.")
    true_curve = convert_to_true_curve(parsed_rows.rows, settings.curve_type)
    elastic_fit = None
    if settings.modulus_mode == "manual":
        if settings.manual_modulus_mpa is None or settings.manual_modulus_mpa <= 0.0:
            raise CurveValidationError("Enter a positive elastic modulus in MPa.")
        modulus_mpa = settings.manual_modulus_mpa
        modulus_origin = "Manual"
    else:
        elastic_fit = infer_elastic_modulus(true_curve)
        modulus_mpa = elastic_fit.modulus_mpa
        modulus_origin = "Inferred"
    miso_rows = build_miso_rows(true_curve, modulus_mpa, settings.proof_offset_plastic_strain)
    summary_parts = [f"Accepted {parsed_rows.accepted_count} curve rows"]
    if parsed_rows.header_skipped:
        summary_parts.append("skipped 1 header row")
    if parsed_rows.skipped_count:
        summary_parts.append(f"skipped {parsed_rows.skipped_count} invalid rows")
    displayed_curve = sorted(
        ((point.input_strain_percent, point.input_stress_mpa) for point in true_curve),
        key=lambda pair: pair[0],
    )
    return ComputationResult(
        miso_rows=miso_rows,
        displayed_curve=displayed_curve,
        modulus_mpa=modulus_mpa,
        modulus_origin=modulus_origin,
        parse_summary=", ".join(summary_parts) + ".",
        elastic_fit=elastic_fit,
    )


def rows_to_tsv(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> str:
    lines = ["\t".join(headers)]
    lines.extend("\t".join(row) for row in rows)
    return "\n".join(lines)


def make_table_item(text: str, editable: bool = True) -> QTableWidgetItem:
    item = QTableWidgetItem(text)
    if not editable:
        item.setFlags(item.flags() & ~ITEM_FLAG.ItemIsEditable)
    return item


class ClipboardTableWidget(QTableWidget):
    paste_requested = pyqtSignal()
    copy_requested = pyqtSignal()

    def keyPressEvent(self, event):  # noqa: N802 - Qt naming
        if event.matches(PASTE_SHORTCUT):
            self.paste_requested.emit()
            event.accept()
            return
        if event.matches(COPY_SHORTCUT):
            self.copy_requested.emit()
            event.accept()
            return
        if event.key() in {KEY_DELETE, KEY_BACKSPACE}:
            for item in self.selectedItems():
                item.setText("")
            event.accept()
            return
        super().keyPressEvent(event)


class CurvePlotWidget(QWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._points: list[tuple[float, float]] = []
        self._highlight_points: list[tuple[float, float]] = []
        self.setMinimumHeight(220)

    def set_points(
        self,
        points: Iterable[tuple[float, float]],
        highlight_points: Iterable[tuple[float, float]] = (),
    ) -> None:
        self._points = list(points)
        self._highlight_points = list(highlight_points)
        self.update()

    def _nice_ticks(self, val_min: float, val_max: float, count: int = 5) -> list[float]:
        span = val_max - val_min
        if span <= 0.0:
            return [val_min]
        raw_step = span / count
        mag = 10.0 ** math.floor(math.log10(raw_step))
        normalized = raw_step / mag
        if normalized < 1.5:
            step = mag
        elif normalized < 3.5:
            step = 2.0 * mag
        elif normalized < 7.5:
            step = 5.0 * mag
        else:
            step = 10.0 * mag
        first = math.ceil(val_min / step) * step
        ticks: list[float] = []
        v = first
        while v <= val_max + step * 0.01:
            ticks.append(round(v, 12))
            v += step
        return ticks

    def paintEvent(self, event):  # noqa: N802 - Qt naming
        del event
        painter = QPainter(self)
        painter.setRenderHint(RENDER_HINT.Antialiasing, True)
        painter.fillRect(self.rect(), QColor("#f7f9fc"))

        margin_left = 72
        margin_right = 18
        margin_top = 16
        margin_bottom = 48
        plot_rect = self.rect().adjusted(margin_left, margin_top, -margin_right, -margin_bottom)
        painter.setPen(QPen(QColor("#b8c2cc"), 1))
        painter.drawRect(plot_rect)

        if len(self._points) < 2:
            painter.setPen(QColor("#4a5568"))
            painter.drawText(
                plot_rect,
                int(ALIGNMENT.AlignCenter),
                "Paste at least two rows to preview the curve.",
            )
            painter.end()
            return

        x_values = [point[0] for point in self._points]
        y_values = [point[1] for point in self._points]
        x_min, x_max = min(x_values), max(x_values)
        y_min, y_max = min(y_values), max(y_values)
        if math.isclose(x_min, x_max):
            x_max = x_min + 1.0
        if math.isclose(y_min, y_max):
            y_max = y_min + 1.0

        x_ticks = self._nice_ticks(x_min, x_max, 5)
        y_ticks = self._nice_ticks(y_min, y_max, 5)

        def map_x(val: float) -> float:
            return plot_rect.left() + (val - x_min) / (x_max - x_min) * plot_rect.width()

        def map_y(val: float) -> float:
            return plot_rect.bottom() - (val - y_min) / (y_max - y_min) * plot_rect.height()

        # Gridlines
        painter.setPen(QPen(QColor("#e2e8f0"), 1))
        for x_tick in x_ticks:
            px = int(map_x(x_tick))
            painter.drawLine(px, plot_rect.top(), px, plot_rect.bottom())
        for y_tick in y_ticks:
            py = int(map_y(y_tick))
            painter.drawLine(plot_rect.left(), py, plot_rect.right(), py)

        # Tick labels
        small_font = painter.font()
        small_font.setPointSize(8)
        painter.setFont(small_font)
        painter.setPen(QColor("#4a5568"))
        for x_tick in x_ticks:
            px = int(map_x(x_tick))
            label = format_number(x_tick, 4)
            painter.drawText(px - 24, plot_rect.bottom() + 4, 48, 14, int(ALIGNMENT.AlignCenter), label)
        for y_tick in y_ticks:
            py = int(map_y(y_tick))
            label = format_number(y_tick, 4)
            painter.drawText(plot_rect.left() - 68, py - 7, 64, 14, int(ALIGNMENT.AlignCenter), label)

        # Axis labels
        axis_font = painter.font()
        axis_font.setPointSize(9)
        painter.setFont(axis_font)
        painter.setPen(QColor("#2d3748"))
        painter.drawText(
            plot_rect.left(),
            self.height() - 14,
            plot_rect.width(),
            14,
            int(ALIGNMENT.AlignCenter),
            "Strain (%)",
        )
        painter.save()
        painter.translate(12, plot_rect.top() + plot_rect.height() // 2)
        painter.rotate(-90)
        painter.drawText(-40, -7, 80, 14, int(ALIGNMENT.AlignCenter), "Stress (MPa)")
        painter.restore()

        # Curve
        painter.setPen(QPen(QColor("#1f5aa6"), 2))
        previous_point = None
        for point in self._points:
            mapped = QPointF(map_x(point[0]), map_y(point[1]))
            if previous_point is not None:
                painter.drawLine(previous_point, mapped)
            previous_point = mapped
        painter.setPen(QPen(QColor("#1f5aa6"), 1))
        painter.setBrush(QColor("#1f5aa6"))
        for point in self._points:
            mapped = QPointF(map_x(point[0]), map_y(point[1]))
            painter.drawEllipse(mapped, 2.5, 2.5)
        if self._highlight_points:
            painter.setPen(QPen(QColor("#d97706"), 1))
            painter.setBrush(QColor("#f59e0b"))
            for point in self._highlight_points:
                mapped = QPointF(map_x(point[0]), map_y(point[1]))
                painter.drawEllipse(mapped, 4.0, 4.0)

        painter.end()


class MisoCurveBuilderWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.temperature_datasets: dict[str, TemperatureCurveData] = {}
        self.temperature_order: list[str] = []
        self.current_temperature_key: str | None = None
        self.current_project_path: Path | None = None
        self._syncing_temperature_table = False
        self._loading_curve_table = False
        self._current_result_rows: list[MisoPoint] = []
        self._current_result_temperature: float | None = None
        self._temperature_status: dict[str, str] = {}
        self._step_num_labels: list[QLabel] = []

        self._update_window_title()
        self.resize(1380, 820)
        self._build_ui()
        self._set_status("Ready. Paste temperatures on the left, select one, then paste stress-strain data.")
        self._recompute_selected_temperature()

    def _update_window_title(self) -> None:
        title = f"MISO Curve Builder ({QT_API})"
        if self.current_project_path is not None:
            title = f"{title} - {self.current_project_path.name}"
        self.setWindowTitle(title)

    def _build_step_banner(self) -> QWidget:
        banner = QWidget()
        layout = QHBoxLayout(banner)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(0)
        steps = [
            "Add Temperatures",
            "Select Temperature",
            "Paste Curve",
            "Configure Settings",
            "Copy / Export",
        ]
        self._step_num_labels = []
        for i, text in enumerate(steps):
            num_label = QLabel(str(i + 1))
            num_label.setFixedSize(22, 22)
            num_label.setAlignment(ALIGNMENT.AlignCenter)
            num_label.setStyleSheet(
                "background: #e5e7eb; color: #6b7280; border-radius: 11px; font-weight: bold;"
            )
            self._step_num_labels.append(num_label)
            text_label = QLabel(text)
            text_label.setStyleSheet("font-size: 11px; color: #374151;")
            step_container = QWidget()
            step_inner = QHBoxLayout(step_container)
            step_inner.setContentsMargins(4, 0, 4, 0)
            step_inner.setSpacing(5)
            step_inner.addWidget(num_label)
            step_inner.addWidget(text_label)
            layout.addWidget(step_container)
            if i < len(steps) - 1:
                arrow = QLabel("→")
                arrow.setStyleSheet("color: #9ca3af; padding: 0 4px;")
                layout.addWidget(arrow)
        layout.addStretch(1)
        banner.setStyleSheet(
            "background: #f9fafb; border: 1px solid #e5e7eb; border-radius: 4px;"
        )
        return banner

    def _refresh_step_banner(self) -> None:
        if not self._step_num_labels:
            return
        if not self.temperature_order:
            active = 0
        elif self.current_temperature_key is None:
            active = 1
        else:
            dataset = self.temperature_datasets.get(self.current_temperature_key)
            has_rows = bool(dataset and dataset.raw_rows)
            if not has_rows:
                active = 2
            else:
                status = self._temperature_status.get(self.current_temperature_key, "")
                active = 4 if status == "ok" else 3
        active_style = (
            "background: #3b82f6; color: white; border-radius: 11px; font-weight: bold;"
        )
        done_style = (
            "background: #10b981; color: white; border-radius: 11px; font-weight: bold;"
        )
        pending_style = (
            "background: #e5e7eb; color: #6b7280; border-radius: 11px; font-weight: bold;"
        )
        for i, label in enumerate(self._step_num_labels):
            if i < active:
                label.setStyleSheet(done_style)
            elif i == active:
                label.setStyleSheet(active_style)
            else:
                label.setStyleSheet(pending_style)

    def _update_temperature_row_color(self, key: str) -> None:
        if key not in self.temperature_order:
            return
        row_index = self.temperature_order.index(key)
        item = self.temperature_table.item(row_index, 0)
        if item is None:
            return
        status = self._temperature_status.get(key, "")
        if status == "ok":
            item.setBackground(QColor("#d1fae5"))
        elif status == "error":
            item.setBackground(QColor("#fee2e2"))
        else:
            item.setBackground(QColor("#ffffff"))

    def _build_ui(self) -> None:
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        self.setToolTip(
            tooltip_html(
                "MISO Curve Builder",
                "Standalone utility for preparing multilinear isotropic hardening input from stress-strain data.",
                "Workflow: paste temperatures, select one temperature, paste its curve, choose conversion settings, then copy or export the generated MISO table.",
            )
        )
        root_layout = QVBoxLayout(central_widget)
        root_layout.setContentsMargins(10, 10, 10, 10)
        root_layout.setSpacing(8)

        guidance_label = QLabel(
            "Material units are fixed: Temperature in C, stress in MPa, strain in percent. "
            "Choose curve type and modulus mode, then the tool converts the selected temperature into ANSYS-ready MISO data."
        )
        guidance_label.setWordWrap(True)
        guidance_label.setToolTip(
            tooltip_html(
                "Workflow Summary",
                "1. Paste one-column temperature data in Celsius on the left.",
                "2. Click a temperature row to make it active.",
                "3. Paste two-column stress-strain data for that temperature in MPa and percent strain.",
                "4. Choose whether the input curve is engineering or true, whether E is manual or inferred, and which proof-stress offset defines yield.",
                "5. Save the work as a project file if you want to resume later.",
                "6. Copy the resulting plastic-strain/stress table or export it to Excel.",
            )
        )
        root_layout.addWidget(guidance_label)
        root_layout.addWidget(self._build_step_banner())

        project_actions = QHBoxLayout()
        self.open_project_button = QPushButton("Open Project")
        self.open_project_button.setToolTip(
            tooltip_html(
                "Open Project",
                "Loads a saved project file and restores temperatures, raw curves, per-temperature manual moduli, the proof-stress selection, and the active conversion settings.",
                "Use this when you want to resume previous work or inspect an old setup.",
            )
        )
        self.open_project_button.clicked.connect(self._open_project)
        project_actions.addWidget(self.open_project_button)

        self.save_project_button = QPushButton("Save Project")
        self.save_project_button.setToolTip(
            tooltip_html(
                "Save Project",
                "Writes the current temperatures, raw curve tables, per-temperature manual moduli, proof-stress selection, and active settings into a project file.",
                "If this session has not been saved before, you will be prompted for a project filename.",
            )
        )
        self.save_project_button.clicked.connect(self._save_project)
        project_actions.addWidget(self.save_project_button)
        project_actions.addStretch(1)
        root_layout.addLayout(project_actions)

        splitter = QSplitter()
        splitter.setToolTip(
            tooltip_html(
                "Main Workspace",
                "The window is split into temperature management, raw curve editing, and computed MISO output.",
                "Drag the splitter handles if you want to resize any panel while keeping the others visible.",
            )
        )
        root_layout.addWidget(splitter, 1)

        splitter.addWidget(self._build_temperature_panel())
        splitter.addWidget(self._build_curve_panel())
        splitter.addWidget(self._build_output_panel())
        splitter.setSizes([240, 560, 560])

        self.validation_label = QLabel("")
        self.validation_label.setWordWrap(True)
        self.validation_label.setStyleSheet("color: #b42318; font-weight: 600;")
        self.validation_label.setToolTip(
            tooltip_html(
                "Validation Message",
                "Shows why conversion or export is currently blocked.",
                "Typical causes are invalid rows, missing elastic modulus, too few curve points, or a curve that never crosses into positive plastic strain.",
            )
        )
        root_layout.addWidget(self.validation_label)

        self.status_box = QTextEdit(self)
        self.status_box.setReadOnly(True)
        self.status_box.setFixedHeight(96)
        self.status_box.setToolTip(
            tooltip_html(
                "Status Log",
                "Records paste summaries, skipped row counts, inferred modulus values, and export status messages.",
                "Use this area to confirm whether a header row or invalid rows were ignored during import.",
            )
        )
        root_layout.addWidget(self.status_box)

    def _build_temperature_panel(self) -> QWidget:
        panel = QGroupBox("Temperatures")
        panel.setToolTip(
            tooltip_html(
                "Temperature List",
                "Stores every temperature for the current material in degrees Celsius.",
                "Each row owns its own stress-strain dataset and computed MISO result.",
                "Duplicate temperatures are removed automatically so each temperature stays unique.",
            )
        )
        layout = QVBoxLayout(panel)

        self.paste_temperatures_button = QPushButton("Paste Temperatures")
        self.paste_temperatures_button.setToolTip(
            tooltip_html(
                "Paste Temperatures",
                "Reads one-column temperature data from the clipboard and inserts it starting at the selected row.",
                "Excel and ANSYS-style clipboard text are both accepted.",
                "Blank lines are ignored, the first nonnumeric row is treated as a header, and later invalid rows are skipped.",
            )
        )
        self.paste_temperatures_button.clicked.connect(self._paste_temperatures_from_clipboard)
        layout.addWidget(self.paste_temperatures_button)

        self.delete_temperature_button = QPushButton("Delete Selected")
        self.delete_temperature_button.setToolTip(
            tooltip_html(
                "Delete Selected Temperature",
                "Removes the currently selected temperature and its associated curve data from the project.",
                "This action cannot be undone within the session.",
            )
        )
        self.delete_temperature_button.clicked.connect(self._delete_selected_temperature)
        layout.addWidget(self.delete_temperature_button)

        self.temperature_table = ClipboardTableWidget(MIN_TEMPERATURE_ROWS, 1, self)
        self.temperature_table.setHorizontalHeaderLabels(["Temperature (C)"])
        self.temperature_table.verticalHeader().setVisible(False)
        self.temperature_table.setSelectionBehavior(SELECT_ROWS)
        self.temperature_table.setSelectionMode(SINGLE_SELECTION)
        self.temperature_table.horizontalHeader().setSectionResizeMode(0, HEADER_STRETCH)
        self.temperature_table.paste_requested.connect(self._paste_temperatures_from_clipboard)
        self.temperature_table.copy_requested.connect(self._copy_selected_temperature)
        self.temperature_table.itemSelectionChanged.connect(self._handle_temperature_selection_change)
        self.temperature_table.itemChanged.connect(self._handle_temperature_table_edit)
        self.temperature_table.setToolTip(
            tooltip_html(
                "Temperature Table",
                "One editable column in Celsius.",
                "Click a row to switch the active temperature and reveal its own curve and result.",
                "Shortcuts: Ctrl+V pastes temperatures, Ctrl+C copies the selected temperature cell, Delete clears selected cells.",
            )
        )
        layout.addWidget(self.temperature_table)
        return panel

    def _build_curve_panel(self) -> QWidget:
        panel = QGroupBox("Raw Curve Input")
        panel.setToolTip(
            tooltip_html(
                "Raw Stress-Strain Input",
                "Contains the curve for the currently selected temperature only.",
                "Units are always Stress = MPa and Strain = percent total strain.",
                "Changing the selected temperature swaps this table to that temperature's saved data.",
            )
        )
        layout = QVBoxLayout(panel)

        controls_layout = QGridLayout()
        curve_type_label = QLabel("Curve Type")
        curve_type_label.setToolTip(
            tooltip_html(
                "Curve Type",
                "Choose how the pasted stress-strain data should be interpreted before MISO conversion.",
                "Engineering: converts stress with sigma_true = sigma_eng * (1 + e_eng) and strain with ln(1 + e_eng).",
                "True: treats the pasted stress as true stress and pasted strain as total true strain in percent.",
            )
        )
        controls_layout.addWidget(curve_type_label, 0, 0)
        self.curve_type_combo = QComboBox(self)
        self.curve_type_combo.addItem("Engineering", "engineering")
        self.curve_type_combo.addItem("True", "true")
        self.curve_type_combo.setToolTip(curve_type_label.toolTip())
        self.curve_type_combo.currentIndexChanged.connect(self._trigger_recompute)
        controls_layout.addWidget(self.curve_type_combo, 0, 1)

        modulus_mode_label = QLabel("Elastic Modulus Mode")
        modulus_mode_label.setToolTip(
            tooltip_html(
                "Elastic Modulus Mode",
                "Controls how the elastic strain portion is removed before building plastic strain.",
                "Manual: uses the E value you type in MPa.",
                "Infer from Curve: estimates E from the initial elastic portion of the converted true curve and displays the fitted value.",
            )
        )
        controls_layout.addWidget(modulus_mode_label, 0, 2)
        self.modulus_mode_combo = QComboBox(self)
        self.modulus_mode_combo.addItem("Manual", "manual")
        self.modulus_mode_combo.addItem("Infer from Curve", "infer")
        self.modulus_mode_combo.setToolTip(modulus_mode_label.toolTip())
        self.modulus_mode_combo.currentIndexChanged.connect(self._handle_modulus_mode_change)
        controls_layout.addWidget(self.modulus_mode_combo, 0, 3)

        proof_offset_label_widget = QLabel("Proof Stress")
        proof_offset_label_widget.setToolTip(
            tooltip_html(
                "Proof-Stress Yield Definition",
                "Controls how the script identifies yield strength before shifting the MISO curve to start at output plastic strain zero.",
                "0.2% proof means yield is taken where computed plastic strain reaches 0.002.",
                "0.02% proof means yield is taken where computed plastic strain reaches 0.0002.",
            )
        )
        controls_layout.addWidget(proof_offset_label_widget, 1, 0)
        self.proof_offset_combo = QComboBox(self)
        self.proof_offset_combo.addItem("0.2% proof (ep = 0.002)", "proof_0p2")
        self.proof_offset_combo.addItem("0.02% proof (ep = 0.0002)", "proof_0p02")
        self.proof_offset_combo.setCurrentIndex(self.proof_offset_combo.findData(DEFAULT_PROOF_OFFSET_MODE))
        self.proof_offset_combo.setToolTip(proof_offset_label_widget.toolTip())
        self.proof_offset_combo.currentIndexChanged.connect(self._trigger_recompute)
        controls_layout.addWidget(self.proof_offset_combo, 1, 1)

        manual_modulus_label = QLabel("E (MPa)")
        manual_modulus_label.setToolTip(
            tooltip_html(
                "Manual Elastic Modulus",
                "Enter Young's modulus in MPa for the currently selected temperature when Elastic Modulus Mode is set to Manual.",
                "The script computes plastic strain as eps_plastic = eps_true - sigma_true / E.",
                "When you switch temperatures, this field shows and stores that temperature's own modulus.",
                "Use a positive value such as 210000 for many steels.",
            )
        )
        controls_layout.addWidget(manual_modulus_label, 2, 0)
        self.manual_modulus_edit = QLineEdit(self)
        self.manual_modulus_edit.setPlaceholderText("Example: 210000")
        self.manual_modulus_edit.setToolTip(manual_modulus_label.toolTip())
        self.manual_modulus_edit.textChanged.connect(self._trigger_recompute)
        controls_layout.addWidget(self.manual_modulus_edit, 2, 1)

        inferred_modulus_label = QLabel("Inferred E (MPa)")
        inferred_modulus_label.setToolTip(
            tooltip_html(
                "Inferred Elastic Modulus",
                "Read-only display of the modulus estimated from the initial elastic segment when Infer from Curve is selected.",
                "The estimate is based on early secant slopes and a line fit over the accepted elastic prefix.",
            )
        )
        controls_layout.addWidget(inferred_modulus_label, 2, 2)
        self.inferred_modulus_value = QLineEdit(self)
        self.inferred_modulus_value.setReadOnly(True)
        self.inferred_modulus_value.setToolTip(inferred_modulus_label.toolTip())
        controls_layout.addWidget(self.inferred_modulus_value, 2, 3)
        fit_summary_label = QLabel("Inference Fit")
        fit_summary_label.setToolTip(
            tooltip_html(
                "Elastic Fit Diagnostics",
                "Summarizes the elastic window used for automatic modulus inference.",
                "Includes the fitted point count, true-strain span, linearity score, and intercept.",
            )
        )
        controls_layout.addWidget(fit_summary_label, 3, 0)
        self.inference_fit_value = QLineEdit(self)
        self.inference_fit_value.setReadOnly(True)
        self.inference_fit_value.setToolTip(fit_summary_label.toolTip())
        controls_layout.addWidget(self.inference_fit_value, 3, 1, 1, 3)
        self.fit_selection_button = QPushButton("Fit Selected Rows to Manual E")
        self.fit_selection_button.setToolTip(
            tooltip_html(
                "Fit Selected Rows",
                "Uses the currently selected stress-strain rows to fit Young's modulus,",
                "writes that value into the manual E field, and switches the active mode to Manual.",
                "Use this when the automatic fit chose the wrong elastic region and you want to lock E from a known-good selection.",
            )
        )
        self.fit_selection_button.clicked.connect(self._fit_selected_rows_to_manual_modulus)
        controls_layout.addWidget(self.fit_selection_button, 4, 0, 1, 4)
        layout.addLayout(controls_layout)

        self.paste_curve_button = QPushButton("Paste Curve")
        self.paste_curve_button.setToolTip(
            tooltip_html(
                "Paste Curve",
                "Reads two-column stress-strain data from the clipboard into the active temperature's table.",
                "The expected order is Stress (MPa), Strain (%).",
                "Blank lines are ignored, the first nonnumeric row is treated as a header, and invalid rows later in the paste are skipped.",
            )
        )
        self.paste_curve_button.clicked.connect(self._paste_curve_from_clipboard)
        layout.addWidget(self.paste_curve_button)

        self.clear_curve_button = QPushButton("Clear Curve")
        self.clear_curve_button.setToolTip(
            tooltip_html(
                "Clear Curve",
                "Removes all stress-strain rows for the currently selected temperature.",
                "The temperature itself is kept; only the curve data is cleared.",
            )
        )
        self.clear_curve_button.clicked.connect(self._clear_current_curve)
        layout.addWidget(self.clear_curve_button)

        self.curve_table = ClipboardTableWidget(MIN_CURVE_ROWS, 2, self)
        self.curve_table.setHorizontalHeaderLabels(["Stress (MPa)", "Strain (%)"])
        self.curve_table.verticalHeader().setVisible(False)
        self.curve_table.setSelectionBehavior(SELECT_ITEMS)
        self.curve_table.setSelectionMode(EXTENDED_SELECTION)
        self.curve_table.horizontalHeader().setSectionResizeMode(0, HEADER_STRETCH)
        self.curve_table.horizontalHeader().setSectionResizeMode(1, HEADER_STRETCH)
        self.curve_table.paste_requested.connect(self._paste_curve_from_clipboard)
        self.curve_table.copy_requested.connect(self._copy_curve_selection)
        self.curve_table.itemChanged.connect(self._handle_curve_table_edit)
        self.curve_table.setToolTip(
            tooltip_html(
                "Stress-Strain Table",
                "Editable two-column table for the currently selected temperature.",
                "Column 1 is stress in MPa. Column 2 is total strain in percent.",
                "Shortcuts: Ctrl+V pastes a two-column block, Ctrl+C copies the current selection, Delete clears selected cells.",
            )
        )
        layout.addWidget(self.curve_table, 1)
        return panel

    def _build_output_panel(self) -> QWidget:
        panel = QWidget(self)
        layout = QVBoxLayout(panel)

        plot_group = QGroupBox("Selected Temperature Preview")
        plot_group.setToolTip(
            tooltip_html(
                "Curve Preview",
                "Plots the raw pasted stress-strain data for the active temperature.",
                "The chart is for visual checking only and uses the input units: strain in percent and stress in MPa.",
                "Changing the selected temperature, editing rows, or switching conversion settings refreshes this preview.",
            )
        )
        plot_layout = QVBoxLayout(plot_group)
        self.plot_widget = CurvePlotWidget(plot_group)
        self.plot_widget.setToolTip(
            tooltip_html(
                "Raw Curve Plot",
                "Visual preview of the pasted stress-strain rows for the active temperature.",
                "This helps confirm that the correct temperature data is selected and that the pasted curve shape looks reasonable before export.",
            )
        )
        plot_layout.addWidget(self.plot_widget)
        layout.addWidget(plot_group)

        result_group = QGroupBox("MISO Output")
        result_group.setToolTip(
            tooltip_html(
                "Multilinear Isotropic Hardening Output",
                "Shows the converted ANSYS-ready table for the active temperature.",
                "The output columns are Plastic Strain (m/m) and Stress (MPa).",
                "The first row is anchored at output plastic strain zero using the selected proof-stress yield point.",
            )
        )
        result_layout = QVBoxLayout(result_group)
        button_row = QHBoxLayout()
        self.copy_result_button = QPushButton("Copy Result")
        self.copy_result_button.setToolTip(
            tooltip_html(
                "Copy Result",
                "Copies the currently displayed MISO table to the clipboard as tab-delimited text with headers.",
                "The copied result can be pasted directly into Excel, Notepad, or ANSYS tabular inputs.",
            )
        )
        self.copy_result_button.clicked.connect(self._copy_result_table)
        button_row.addWidget(self.copy_result_button)
        self.export_selected_button = QPushButton("Export Selected")
        self.export_selected_button.setToolTip(
            tooltip_html(
                "Export Selected",
                "Writes only the active temperature's MISO output to a new Excel workbook.",
                "The workbook contains one worksheet named from the selected temperature.",
            )
        )
        self.export_selected_button.clicked.connect(self._export_selected_temperature)
        button_row.addWidget(self.export_selected_button)
        self.export_all_button = QPushButton("Export All")
        self.export_all_button.setToolTip(
            tooltip_html(
                "Export All",
                "Builds one Excel workbook with one worksheet per valid temperature.",
                "Temperatures that cannot be converted are skipped and reported after export.",
            )
        )
        self.export_all_button.clicked.connect(self._export_all_temperatures)
        button_row.addWidget(self.export_all_button)
        button_row.addStretch(1)
        result_layout.addLayout(button_row)

        self.result_table = ClipboardTableWidget(0, 2, self)
        self.result_table.setHorizontalHeaderLabels(["Plastic Strain (m/m)", "Stress (MPa)"])
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setEditTriggers(NO_EDIT_TRIGGERS)
        self.result_table.setSelectionBehavior(SELECT_ITEMS)
        self.result_table.setSelectionMode(EXTENDED_SELECTION)
        self.result_table.horizontalHeader().setSectionResizeMode(0, HEADER_STRETCH)
        self.result_table.horizontalHeader().setSectionResizeMode(1, HEADER_STRETCH)
        self.result_table.copy_requested.connect(self._copy_result_table)
        self.result_table.setToolTip(
            tooltip_html(
                "MISO Result Table",
                "Read-only output table for the active temperature.",
                "Column 1 is plastic strain in m/m. Column 2 is stress in MPa.",
                "Use Ctrl+C or the Copy Result button to copy the displayed rows with headers.",
            )
        )
        result_layout.addWidget(self.result_table)
        layout.addWidget(result_group, 1)
        return panel

    def _set_status(self, message: object) -> None:
        if message is None:
            return
        self.status_box.append(str(message))
        cursor = self.status_box.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        self.status_box.setTextCursor(cursor)

    def _show_validation_error(self, message: str) -> None:
        self.validation_label.setText(message)

    def _clear_validation_error(self) -> None:
        self.validation_label.setText("")

    def _copy_text_to_clipboard(self, text: str) -> None:
        QApplication.clipboard().setText(text)

    def _project_payload(self) -> dict:
        self._save_current_curve_to_dataset()
        self._save_current_manual_modulus_to_dataset()
        return {
            "version": PROJECT_FILE_VERSION,
            "curve_type": self.curve_type_combo.currentData(),
            "modulus_mode": self.modulus_mode_combo.currentData(),
            "proof_offset_mode": self.proof_offset_combo.currentData(),
            "current_temperature_key": self.current_temperature_key,
            "temperatures": [
                {
                    "temperature_c": self.temperature_datasets[key].temperature_c,
                    "manual_modulus_text": self.temperature_datasets[key].manual_modulus_text,
                    "raw_rows": [
                        {
                            "stress_text": row.stress_text,
                            "strain_text": row.strain_text,
                        }
                        for row in self.temperature_datasets[key].raw_rows
                    ],
                }
                for key in self.temperature_order
            ],
        }

    def _save_project_to_path(self, path: Path) -> None:
        payload = self._project_payload()
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def _load_project_from_path(self, path: Path) -> None:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except OSError as exc:
            raise ProjectFileError(f"Could not read project file: {exc}") from exc
        except json.JSONDecodeError as exc:
            raise ProjectFileError(f"Project file is not valid JSON: {exc}") from exc

        if not isinstance(payload, dict):
            raise ProjectFileError("Project file root must be a JSON object.")

        temperatures_payload = payload.get("temperatures")
        if not isinstance(temperatures_payload, list):
            raise ProjectFileError("Project file is missing a valid temperatures list.")

        curve_type = payload.get("curve_type", "engineering")
        modulus_mode = payload.get("modulus_mode", "manual")
        proof_offset_mode = payload.get("proof_offset_mode", DEFAULT_PROOF_OFFSET_MODE)
        if curve_type not in {"engineering", "true"}:
            raise ProjectFileError("Project file contains an unsupported curve type.")
        if modulus_mode not in {"manual", "infer"}:
            raise ProjectFileError("Project file contains an unsupported modulus mode.")
        if proof_offset_mode not in {"proof_0p2", "proof_0p02"}:
            raise ProjectFileError("Project file contains an unsupported proof-stress option.")

        datasets: dict[str, TemperatureCurveData] = {}
        order: list[str] = []
        for entry in temperatures_payload:
            if not isinstance(entry, dict):
                raise ProjectFileError("Each saved temperature entry must be a JSON object.")
            try:
                temperature_c = float(entry["temperature_c"])
            except (KeyError, TypeError, ValueError) as exc:
                raise ProjectFileError("A saved temperature entry is missing a valid temperature value.") from exc
            key = temperature_key(temperature_c)
            if key in datasets:
                raise ProjectFileError("Project file contains duplicate temperatures.")
            raw_rows_payload = entry.get("raw_rows", [])
            if not isinstance(raw_rows_payload, list):
                raise ProjectFileError("Saved raw_rows must be a list.")
            raw_rows = [
                RawCurveRow(
                    stress_text=str(row.get("stress_text", "")),
                    strain_text=str(row.get("strain_text", "")),
                )
                for row in raw_rows_payload
                if isinstance(row, dict)
            ]
            datasets[key] = TemperatureCurveData(
                temperature_c=temperature_c,
                raw_rows=raw_rows,
                manual_modulus_text=str(entry.get("manual_modulus_text", "")),
            )
            order.append(key)

        requested_key = payload.get("current_temperature_key")
        current_key = requested_key if requested_key in datasets else (order[0] if order else None)

        with QSignalBlocker(self.curve_type_combo):
            self.curve_type_combo.setCurrentIndex(self.curve_type_combo.findData(curve_type))
        with QSignalBlocker(self.modulus_mode_combo):
            self.modulus_mode_combo.setCurrentIndex(self.modulus_mode_combo.findData(modulus_mode))
        with QSignalBlocker(self.proof_offset_combo):
            self.proof_offset_combo.setCurrentIndex(self.proof_offset_combo.findData(proof_offset_mode))

        self.temperature_datasets = datasets
        self.temperature_order = order
        self.current_temperature_key = current_key
        self.current_project_path = path
        self._update_window_title()
        self._refresh_temperature_table()
        self._load_curve_for_temperature(current_key)
        self.manual_modulus_edit.setEnabled(self.modulus_mode_combo.currentData() == "manual")
        self._recompute_selected_temperature(extra_status=f"Loaded project from {path}.")

    def _save_project(self) -> None:
        target_path = self.current_project_path
        if target_path is None:
            path_text, _ = QFileDialog.getSaveFileName(
                self,
                "Save Project",
                str(Path.cwd() / "miso_project.miso.json"),
                "MISO Project (*.miso.json);;JSON Files (*.json)",
            )
            if not path_text:
                return
            target_path = Path(path_text)
        try:
            self._save_project_to_path(target_path)
        except OSError as exc:
            QMessageBox.critical(self, "Save Project Failed", f"Could not save project:\n\n{exc}")
            return
        self.current_project_path = target_path
        self._update_window_title()
        self._set_status(f"Saved project to {target_path}.")

    def _open_project(self) -> None:
        path_text, _ = QFileDialog.getOpenFileName(
            self,
            "Open Project",
            str(Path.cwd()),
            "MISO Project (*.miso.json *.json);;JSON Files (*.json)",
        )
        if not path_text:
            return
        try:
            self._load_project_from_path(Path(path_text))
        except ProjectFileError as exc:
            QMessageBox.critical(self, "Open Project Failed", str(exc))

    def _parse_manual_modulus_text(self, modulus_text: str) -> float | None:
        manual_modulus = None
        if modulus_text:
            try:
                manual_modulus = parse_localized_float(modulus_text)
            except ValueError:
                manual_modulus = None
        return manual_modulus

    def _settings_for_dataset(self, dataset: TemperatureCurveData | None) -> ConversionSettings:
        modulus_text = dataset.manual_modulus_text.strip() if dataset is not None else ""
        proof_mode = self.proof_offset_combo.currentData() or DEFAULT_PROOF_OFFSET_MODE
        return ConversionSettings(
            curve_type=self.curve_type_combo.currentData(),
            modulus_mode=self.modulus_mode_combo.currentData(),
            manual_modulus_mpa=self._parse_manual_modulus_text(modulus_text),
            proof_offset_mode=proof_mode,
            proof_offset_plastic_strain=proof_offset_value(proof_mode),
        )

    def _trigger_recompute(self, *_args) -> None:
        self._save_current_manual_modulus_to_dataset()
        self._recompute_selected_temperature()

    def _handle_modulus_mode_change(self, *_args) -> None:
        is_manual = self.modulus_mode_combo.currentData() == "manual"
        self.manual_modulus_edit.setEnabled(is_manual)
        self._recompute_selected_temperature()

    def _copy_selected_temperature(self) -> None:
        selected_row = self.temperature_table.currentRow()
        if selected_row < 0:
            return
        item = self.temperature_table.item(selected_row, 0)
        if item is None:
            return
        self._copy_text_to_clipboard(item.text())

    def _copy_curve_selection(self) -> None:
        selected_items = self.curve_table.selectedItems()
        if not selected_items:
            return
        indexed = {(item.row(), item.column()): item.text() for item in selected_items}
        min_row = min(row for row, _ in indexed)
        max_row = max(row for row, _ in indexed)
        min_col = min(col for _, col in indexed)
        max_col = max(col for _, col in indexed)
        rows: list[list[str]] = []
        for row_index in range(min_row, max_row + 1):
            row_values = []
            for column_index in range(min_col, max_col + 1):
                row_values.append(indexed.get((row_index, column_index), ""))
            rows.append(row_values)
        self._copy_text_to_clipboard("\n".join("\t".join(row) for row in rows))

    def _fit_selected_rows_to_manual_modulus(self) -> None:
        selected_row_indexes = sorted({item.row() for item in self.curve_table.selectedItems()})
        if len(selected_row_indexes) < 2:
            QMessageBox.information(
                self,
                "Fit Selected Rows",
                "Select at least two stress-strain rows before fitting Young's modulus.",
            )
            return
        selected_rows: list[RawCurveRow] = []
        for row_index in selected_row_indexes:
            stress_item = self.curve_table.item(row_index, 0)
            strain_item = self.curve_table.item(row_index, 1)
            selected_rows.append(
                RawCurveRow(
                    stress_text=stress_item.text().strip() if stress_item else "",
                    strain_text=strain_item.text().strip() if strain_item else "",
                )
            )
        parsed_rows = parse_curve_rows(selected_rows)
        if parsed_rows.accepted_count < 2:
            QMessageBox.warning(
                self,
                "Fit Selected Rows",
                "The selected rows do not contain at least two valid stress-strain pairs.",
            )
            return
        try:
            fit = infer_elastic_modulus(convert_to_true_curve(parsed_rows.rows, self.curve_type_combo.currentData()))
        except CurveValidationError as exc:
            QMessageBox.warning(self, "Fit Selected Rows", str(exc))
            return
        self.manual_modulus_edit.setText(format_number(fit.modulus_mpa, 12))
        self.modulus_mode_combo.setCurrentIndex(self.modulus_mode_combo.findData("manual"))
        self._set_status(f"Locked manual E from selected rows. {fit_summary_text(fit)}.")

    def _copy_result_table(self) -> None:
        if not self._current_result_rows:
            QMessageBox.information(self, "No Result", "No computed MISO data is available to copy.")
            return
        tsv = rows_to_tsv(
            headers=["Plastic Strain (m/m)", "Stress (MPa)"],
            rows=[
                [format_number(row.plastic_strain, 12), format_number(row.stress_mpa, 12)]
                for row in self._current_result_rows
            ],
        )
        self._copy_text_to_clipboard(tsv)
        self._set_status("Copied the selected temperature result table to the clipboard.")

    def _delete_selected_temperature(self) -> None:
        key = self.current_temperature_key
        if key is None:
            QMessageBox.warning(self, "Delete Temperature", "No temperature is selected.")
            return
        dataset = self.temperature_datasets.get(key)
        label = format_temperature_label(dataset.temperature_c) if dataset else key
        self.temperature_datasets.pop(key, None)
        if key in self.temperature_order:
            self.temperature_order.remove(key)
        self._temperature_status.pop(key, None)
        new_key = self.temperature_order[0] if self.temperature_order else None
        self.current_temperature_key = new_key
        self._refresh_temperature_table()
        self._load_curve_for_temperature(new_key)
        self._recompute_selected_temperature(extra_status=f"Deleted temperature {label}.")

    def _clear_current_curve(self) -> None:
        if self.current_temperature_key is None:
            QMessageBox.warning(self, "Clear Curve", "No temperature is selected.")
            return
        dataset = self.temperature_datasets.get(self.current_temperature_key)
        if dataset:
            dataset.raw_rows = []
        self._load_curve_for_temperature(self.current_temperature_key)
        self._recompute_selected_temperature(extra_status="Cleared curve for selected temperature.")

    def _paste_temperatures_from_clipboard(self) -> None:
        clipboard_text = QApplication.clipboard().text()
        parsed = parse_numeric_clipboard(clipboard_text, expected_columns=1)
        if not parsed.rows:
            QMessageBox.warning(self, "Paste Temperatures", "Clipboard does not contain any valid temperature values.")
            return
        anchor_row = self.temperature_table.currentRow()
        if anchor_row < 0:
            anchor_row = 0
        required_rows = max(self.temperature_table.rowCount(), anchor_row + len(parsed.rows), MIN_TEMPERATURE_ROWS)
        with QSignalBlocker(self.temperature_table):
            self.temperature_table.setRowCount(required_rows)
            for offset, row_values in enumerate(parsed.rows):
                self.temperature_table.setItem(anchor_row + offset, 0, QTableWidgetItem(row_values[0]))
        self._sync_temperatures_from_table(
            f"Pasted {parsed.accepted_count} temperature rows"
            + (" and skipped 1 header row" if parsed.header_skipped else "")
            + (f"; skipped {parsed.skipped_count} invalid rows." if parsed.skipped_count else ".")
        )

    def _paste_curve_from_clipboard(self) -> None:
        if self.current_temperature_key is None:
            QMessageBox.warning(self, "Paste Curve", "Paste or select a temperature before adding a curve.")
            return
        clipboard_text = QApplication.clipboard().text()
        parsed = parse_numeric_clipboard(clipboard_text, expected_columns=2)
        if not parsed.rows:
            QMessageBox.warning(self, "Paste Curve", "Clipboard does not contain valid two-column stress-strain data.")
            return
        anchor_row = self.curve_table.currentRow()
        if anchor_row < 0:
            anchor_row = 0
        required_rows = max(self.curve_table.rowCount(), anchor_row + len(parsed.rows), MIN_CURVE_ROWS)
        self._loading_curve_table = True
        with QSignalBlocker(self.curve_table):
            self.curve_table.setRowCount(required_rows)
            for offset, row_values in enumerate(parsed.rows):
                self.curve_table.setItem(anchor_row + offset, 0, QTableWidgetItem(row_values[0]))
                self.curve_table.setItem(anchor_row + offset, 1, QTableWidgetItem(row_values[1]))
        self._loading_curve_table = False
        self._save_current_curve_to_dataset()
        self._recompute_selected_temperature(
            extra_status=(
                f"Pasted {parsed.accepted_count} curve rows"
                + (" and skipped 1 header row" if parsed.header_skipped else "")
                + (f"; skipped {parsed.skipped_count} invalid rows." if parsed.skipped_count else ".")
            )
        )

    def _handle_temperature_table_edit(self, item: QTableWidgetItem) -> None:
        if self._syncing_temperature_table:
            return
        del item
        self._sync_temperatures_from_table("Updated temperature table.")

    def _handle_curve_table_edit(self, item: QTableWidgetItem) -> None:
        if self._loading_curve_table:
            return
        del item
        self._save_current_curve_to_dataset()
        self._recompute_selected_temperature()

    def _save_current_manual_modulus_to_dataset(self) -> None:
        if self.current_temperature_key is None:
            return
        dataset = self.temperature_datasets.get(self.current_temperature_key)
        if dataset is None:
            return
        dataset.manual_modulus_text = self.manual_modulus_edit.text().strip()

    def _save_current_curve_to_dataset(self) -> None:
        if self.current_temperature_key is None:
            return
        dataset = self.temperature_datasets.get(self.current_temperature_key)
        if dataset is None:
            return
        rows: list[RawCurveRow] = []
        for row_index in range(self.curve_table.rowCount()):
            stress_item = self.curve_table.item(row_index, 0)
            strain_item = self.curve_table.item(row_index, 1)
            stress_text = stress_item.text().strip() if stress_item else ""
            strain_text = strain_item.text().strip() if strain_item else ""
            if stress_text or strain_text:
                rows.append(RawCurveRow(stress_text=stress_text, strain_text=strain_text))
        dataset.raw_rows = rows

    def _load_manual_modulus_for_temperature(self, key: str | None) -> None:
        dataset = self.temperature_datasets.get(key) if key else None
        modulus_text = dataset.manual_modulus_text if dataset is not None else ""
        with QSignalBlocker(self.manual_modulus_edit):
            self.manual_modulus_edit.setText(modulus_text)

    def _load_curve_for_temperature(self, key: str | None) -> None:
        self._loading_curve_table = True
        with QSignalBlocker(self.curve_table):
            row_count = MIN_CURVE_ROWS
            dataset = self.temperature_datasets.get(key) if key else None
            if dataset:
                row_count = max(MIN_CURVE_ROWS, len(dataset.raw_rows))
            self.curve_table.setRowCount(row_count)
            for row_index in range(row_count):
                for column_index in range(2):
                    self.curve_table.takeItem(row_index, column_index)
            if dataset:
                for row_index, row in enumerate(dataset.raw_rows):
                    self.curve_table.setItem(row_index, 0, QTableWidgetItem(row.stress_text))
                    self.curve_table.setItem(row_index, 1, QTableWidgetItem(row.strain_text))
        self._loading_curve_table = False
        self._load_manual_modulus_for_temperature(key)

    def _refresh_temperature_table(self) -> None:
        self._syncing_temperature_table = True
        with QSignalBlocker(self.temperature_table):
            row_count = max(MIN_TEMPERATURE_ROWS, len(self.temperature_order))
            self.temperature_table.setRowCount(row_count)
            for row_index in range(row_count):
                self.temperature_table.takeItem(row_index, 0)
            for row_index, key in enumerate(self.temperature_order):
                self.temperature_table.setItem(
                    row_index,
                    0,
                    QTableWidgetItem(format_number(self.temperature_datasets[key].temperature_c, 9)),
                )
                self._update_temperature_row_color(key)
            if self.current_temperature_key is not None and self.current_temperature_key in self.temperature_order:
                self.temperature_table.selectRow(self.temperature_order.index(self.current_temperature_key))
            else:
                self.temperature_table.clearSelection()
        self._syncing_temperature_table = False

    def _sync_temperatures_from_table(self, status_message: str) -> None:
        self._save_current_curve_to_dataset()
        self._save_current_manual_modulus_to_dataset()
        raw_values = []
        for row_index in range(self.temperature_table.rowCount()):
            item = self.temperature_table.item(row_index, 0)
            raw_values.append(item.text() if item else "")
        parsed = parse_temperature_table_rows(raw_values)
        existing_by_key = self.temperature_datasets
        new_datasets: dict[str, TemperatureCurveData] = {}
        new_order: list[str] = []
        duplicate_count = 0
        for row_values in parsed.rows:
            temperature = parse_localized_float(row_values[0])
            key = temperature_key(temperature)
            if key in new_datasets:
                duplicate_count += 1
                continue
            existing = existing_by_key.get(key)
            if existing:
                existing.temperature_c = temperature
                new_datasets[key] = existing
            else:
                new_datasets[key] = TemperatureCurveData(temperature_c=temperature)
            new_order.append(key)

        current_key = self.current_temperature_key if self.current_temperature_key in new_datasets else None
        if current_key is None and new_order:
            current_key = new_order[0]

        self.temperature_datasets = new_datasets
        self.temperature_order = new_order
        self.current_temperature_key = current_key

        self._refresh_temperature_table()
        self._load_curve_for_temperature(current_key)
        self._recompute_selected_temperature(
            extra_status=(
                status_message
                + (f" Removed {duplicate_count} duplicate temperatures." if duplicate_count else "")
                + (" Skipped 1 header row." if parsed.header_skipped else "")
                + (f" Skipped {parsed.skipped_count} invalid temperature rows." if parsed.skipped_count else "")
            )
        )

    def _handle_temperature_selection_change(self) -> None:
        if self._syncing_temperature_table:
            return
        selected_row = self.temperature_table.currentRow()
        new_key = None
        if 0 <= selected_row < len(self.temperature_order):
            new_key = self.temperature_order[selected_row]
        if new_key == self.current_temperature_key:
            return
        self._save_current_curve_to_dataset()
        self._save_current_manual_modulus_to_dataset()
        self.current_temperature_key = new_key
        self._load_curve_for_temperature(new_key)
        self._recompute_selected_temperature()

    def _display_result_rows(self, rows: Sequence[MisoPoint]) -> None:
        with QSignalBlocker(self.result_table):
            self.result_table.setRowCount(len(rows))
            for row_index, row in enumerate(rows):
                self.result_table.setItem(
                    row_index,
                    0,
                    make_table_item(format_number(row.plastic_strain, 12), editable=False),
                )
                self.result_table.setItem(
                    row_index,
                    1,
                    make_table_item(format_number(row.stress_mpa, 12), editable=False),
                )

    def _clear_result_display(self) -> None:
        self._current_result_rows = []
        self._current_result_temperature = None
        self.inferred_modulus_value.clear()
        self.inference_fit_value.clear()
        self.plot_widget.set_points([], [])
        self._display_result_rows([])

    def _recompute_selected_temperature(self, extra_status: str | None = None) -> None:
        self._clear_validation_error()
        if self.current_temperature_key is None:
            self._clear_result_display()
            if extra_status:
                self._set_status(extra_status)
            self._refresh_step_banner()
            return
        dataset = self.temperature_datasets.get(self.current_temperature_key)
        if dataset is None:
            self._clear_result_display()
            if extra_status:
                self._set_status(extra_status)
            self._refresh_step_banner()
            return
        self._save_current_curve_to_dataset()
        self._save_current_manual_modulus_to_dataset()
        settings = self._settings_for_dataset(dataset)
        try:
            result = compute_miso_from_rows(dataset.raw_rows, settings)
        except CurveValidationError as exc:
            self._temperature_status[self.current_temperature_key] = "error"
            self._update_temperature_row_color(self.current_temperature_key)
            self._clear_result_display()
            self._show_validation_error(str(exc))
            if extra_status:
                self._set_status(extra_status)
            self._refresh_step_banner()
            return

        self._current_result_rows = result.miso_rows
        self._current_result_temperature = dataset.temperature_c
        if result.modulus_origin == "Inferred":
            self.inferred_modulus_value.setText(format_number(result.modulus_mpa, 12))
            fit = result.elastic_fit
            self.inference_fit_value.setText(fit_summary_text(fit) if fit is not None else "")
        else:
            self.inferred_modulus_value.clear()
            self.inference_fit_value.clear()
        highlighted_curve = []
        if result.elastic_fit is not None:
            highlighted_curve = [
                (point.input_strain_percent, point.input_stress_mpa) for point in result.elastic_fit.points
            ]
        self.plot_widget.set_points(result.displayed_curve, highlighted_curve)
        self._display_result_rows(result.miso_rows)

        status_parts = [
            result.parse_summary,
            f"Yield rule = {proof_offset_label(settings.proof_offset_mode)} proof.",
            f"{result.modulus_origin} E = {format_number(result.modulus_mpa, 12)} MPa.",
        ]
        if result.elastic_fit is not None:
            status_parts.append(f"Elastic fit: {fit_summary_text(result.elastic_fit)}.")
        if extra_status:
            status_parts.insert(0, extra_status)
        self._set_status(" ".join(status_parts))
        self._temperature_status[self.current_temperature_key] = "ok"
        self._update_temperature_row_color(self.current_temperature_key)
        self._refresh_step_banner()

    def _collect_result_rows_for_temperature(self, key: str) -> list[MisoPoint]:
        dataset = self.temperature_datasets[key]
        result = compute_miso_from_rows(dataset.raw_rows, self._settings_for_dataset(dataset))
        return result.miso_rows

    def _write_sheet(self, workbook: Workbook, title: str, rows: Sequence[MisoPoint], used_names: set[str]) -> None:
        sheet = workbook.create_sheet(title=sanitize_sheet_name(title, used_names))
        sheet.append(["Plastic Strain (m/m)", "Stress (MPa)"])
        for row in rows:
            sheet.append([row.plastic_strain, row.stress_mpa])
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 18

    def _export_selected_temperature(self) -> None:
        if self.current_temperature_key is None or not self._current_result_rows:
            QMessageBox.warning(self, "Export Selected", "No selected temperature result is available to export.")
            return
        if Workbook is None:
            QMessageBox.critical(self, "Missing Dependency", "openpyxl is required for Excel export.")
            return
        dataset = self.temperature_datasets[self.current_temperature_key]
        default_name = f"miso_{format_number(dataset.temperature_c, 6)}C.xlsx".replace(" ", "_")
        target_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Selected Temperature",
            str(Path.cwd() / default_name),
            "Excel Workbook (*.xlsx)",
        )
        if not target_path:
            return
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        used_names: set[str] = set()
        self._write_sheet(
            workbook,
            title=f"T_{format_number(dataset.temperature_c, 8)}C",
            rows=self._current_result_rows,
            used_names=used_names,
        )
        workbook.save(target_path)
        self._set_status(f"Exported selected temperature result to {target_path}.")

    def _export_all_temperatures(self) -> None:
        if not self.temperature_order:
            QMessageBox.warning(self, "Export All", "No temperatures are available to export.")
            return
        if Workbook is None:
            QMessageBox.critical(self, "Missing Dependency", "openpyxl is required for Excel export.")
            return
        target_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export All Temperatures",
            str(Path.cwd() / "miso_all_temperatures.xlsx"),
            "Excel Workbook (*.xlsx)",
        )
        if not target_path:
            return
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        used_names: set[str] = set()
        failures: list[str] = []
        for key in self.temperature_order:
            dataset = self.temperature_datasets[key]
            try:
                rows = self._collect_result_rows_for_temperature(key)
            except CurveValidationError as exc:
                failures.append(f"{format_temperature_label(dataset.temperature_c)}: {exc}")
                continue
            self._write_sheet(
                workbook,
                title=f"T_{format_number(dataset.temperature_c, 8)}C",
                rows=rows,
                used_names=used_names,
            )
        if not workbook.sheetnames:
            QMessageBox.warning(
                self,
                "Export All",
                "No temperature had valid MISO output to export.\n\n" + "\n".join(failures[:10]),
            )
            return
        workbook.save(target_path)
        if failures:
            QMessageBox.warning(
                self,
                "Export All Completed With Skips",
                "Export finished, but some temperatures were skipped:\n\n" + "\n".join(failures[:10]),
            )
        self._set_status(f"Exported all valid temperatures to {target_path}.")


def main() -> int:
    app = QApplication(sys.argv)
    window = MisoCurveBuilderWindow()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
