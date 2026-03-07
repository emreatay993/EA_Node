from __future__ import annotations

import csv
import json
import os
import queue
import shlex
import smtplib
import subprocess
import sys
import threading
import time
from collections import deque
from email.message import EmailMessage
from pathlib import Path
from typing import Any

from ea_node_editor.nodes.types import (
    ExecutionContext,
    NodeResult,
    NodeTypeSpec,
    PortSpec,
    PropertySpec,
)

try:
    import openpyxl  # type: ignore
except Exception:  # noqa: BLE001
    openpyxl = None


_PROCESS_STREAM_CAPTURE_CHAR_LIMIT = 262_144
_PROCESS_STREAM_QUEUE_SIZE = 256


def _pick_path(ctx: ExecutionContext, *, input_key: str, property_key: str, node_name: str) -> Path:
    for candidate in (ctx.inputs.get(input_key), ctx.properties.get(property_key)):
        if candidate is None:
            continue
        text = str(candidate).strip()
        if text:
            return Path(text)
    raise ValueError(f"{node_name} requires a non-empty file path.")


def _require_existing_file(path: Path, *, node_name: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{node_name} path does not exist: {path}")
    if not path.is_file():
        raise ValueError(f"{node_name} path must point to a file: {path}")


def _require_openpyxl(*, node_name: str) -> None:
    if openpyxl is None:
        runtime_mode = "packaged" if bool(getattr(sys, "frozen", False)) else "source"
        install_guidance = (
            "Rebuild package with openpyxl installed in the build environment."
            if runtime_mode == "packaged"
            else "Install with: pip install openpyxl"
        )
        raise RuntimeError(
            f"{node_name} requires optional dependency 'openpyxl' for XLSX support. "
            f"Runtime mode: {runtime_mode}. CSV remains supported without this dependency. "
            f"{install_guidance}"
        )


def _normalize_headers(values: tuple[Any, ...] | list[Any]) -> list[str]:
    counters: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values):
        base = str(value).strip() if value is not None else ""
        if not base:
            base = f"column_{index + 1}"
        count = counters.get(base, 0) + 1
        counters[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _normalize_rows_input(rows_input: Any) -> list[dict[str, Any]]:
    if not isinstance(rows_input, list) or not rows_input:
        raise ValueError("Excel Write requires 'rows' as a non-empty list of dictionaries.")
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows_input, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"Excel Write row {index} must be a dictionary.")
        rows.append({str(key): value for key, value in row.items()})
    return rows


def _stable_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers = sorted({key for row in rows for key in row})
    if not headers:
        raise ValueError("Excel Write rows must contain at least one column key.")
    return headers


def _split_recipients(value: str) -> list[str]:
    normalized = value.replace(";", ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]


def _normalize_args(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    text = str(value).strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [str(item) for item in parsed]
    except Exception:  # noqa: BLE001
        pass
    return shlex.split(text, posix=False)


def _normalize_env(value: Any) -> dict[str, str]:
    if isinstance(value, dict):
        return {str(key): str(item) for key, item in value.items()}
    text = str(value).strip()
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except Exception:  # noqa: BLE001
        return {}
    if isinstance(parsed, dict):
        return {str(key): str(item) for key, item in parsed.items()}
    return {}


class ExcelReadNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.excel_read",
            display_name="Excel Read",
            category="Input / Output",
            icon="table_view",
            description="Loads rows from Excel or CSV.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("path", "in", "data", "path", required=False),
                PortSpec("rows", "out", "data", "list[dict]", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
            ),
            properties=(
                PropertySpec("path", "path", "", "File Path"),
                PropertySpec("sheet_name", "str", "", "Sheet Name"),
            ),
        )

    def execute(self, ctx: ExecutionContext) -> NodeResult:
        path = _pick_path(ctx, input_key="path", property_key="path", node_name="Excel Read")
        _require_existing_file(path, node_name="Excel Read")
        suffix = path.suffix.lower()

        rows: list[dict[str, Any]] = []
        if suffix == ".csv":
            with path.open("r", encoding="utf-8", newline="") as handle:
                reader = csv.reader(handle)
                header_row = next(reader, None)
                if header_row is not None:
                    headers = _normalize_headers(header_row)
                    for source_row in reader:
                        rows.append(
                            {
                                headers[index]: source_row[index] if index < len(source_row) else ""
                                for index in range(len(headers))
                            }
                        )
        elif suffix in {".xlsx", ".xlsm"}:
            _require_openpyxl(node_name="Excel Read")
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
            sheet_name = str(ctx.properties.get("sheet_name", "")).strip()
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    available = ", ".join(workbook.sheetnames)
                    raise ValueError(
                        f"Excel Read sheet '{sheet_name}' was not found. Available sheets: {available}"
                    )
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            try:
                iterator = sheet.iter_rows(values_only=True)
                header_row = next(iterator, None)
                if header_row is not None:
                    headers = _normalize_headers(list(header_row))
                    for source_row in iterator:
                        values = list(source_row)
                        rows.append(
                            {
                                headers[index]: values[index] if index < len(values) else None
                                for index in range(len(headers))
                            }
                        )
            finally:
                workbook.close()
        else:
            raise ValueError(
                "Excel Read supports only .csv, .xlsx, and .xlsm files. "
                f"Received: {path.suffix or '<no extension>'}"
            )
        return NodeResult(outputs={"rows": rows, "exec_out": True})


class ExcelWriteNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.excel_write",
            display_name="Excel Write",
            category="Input / Output",
            icon="download",
            description="Writes list[dict] rows to CSV/XLSX.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("rows", "in", "data", "list[dict]", required=True),
                PortSpec("path", "in", "data", "path", required=False),
                PortSpec("written_path", "out", "data", "path", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
            ),
            properties=(
                PropertySpec("path", "path", "output.csv", "Output Path"),
            ),
        )

    def execute(self, ctx: ExecutionContext) -> NodeResult:
        rows = _normalize_rows_input(ctx.inputs.get("rows"))
        headers = _stable_headers(rows)
        path = _pick_path(ctx, input_key="path", property_key="path", node_name="Excel Write")
        suffix = path.suffix.lower()
        if suffix not in {".csv", ".xlsx", ".xlsm"}:
            raise ValueError(
                "Excel Write supports only .csv, .xlsx, and .xlsm output formats. "
                f"Received: {path.suffix or '<no extension>'}"
            )
        if path.exists() and path.is_dir():
            raise ValueError(f"Excel Write path must be a file, not a directory: {path}")

        path.parent.mkdir(parents=True, exist_ok=True)
        if suffix == ".csv":
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                writer.writeheader()
                for row in rows:
                    writer.writerow({header: row.get(header) for header in headers})
        else:
            _require_openpyxl(node_name="Excel Write")
            workbook = openpyxl.Workbook()  # type: ignore[union-attr]
            try:
                sheet = workbook.active
                sheet.append(headers)
                for row in rows:
                    sheet.append([row.get(header) for header in headers])
                workbook.save(path)
            finally:
                workbook.close()
        return NodeResult(outputs={"written_path": str(path), "exec_out": True})


class FileReadNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.file_read",
            display_name="File Read",
            category="Input / Output",
            icon="description",
            description="Reads a text file.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("path", "in", "data", "path", required=False),
                PortSpec("text", "out", "data", "str", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
            ),
            properties=(PropertySpec("path", "path", "", "File Path"),),
        )

    def execute(self, ctx: ExecutionContext) -> NodeResult:
        path = _pick_path(ctx, input_key="path", property_key="path", node_name="File Read")
        _require_existing_file(path, node_name="File Read")
        try:
            text = path.read_text(encoding="utf-8")
        except OSError as exc:
            raise RuntimeError(f"File Read failed for '{path}': {exc}") from exc
        return NodeResult(outputs={"text": text, "exec_out": True})


class FileWriteNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.file_write",
            display_name="File Write",
            category="Input / Output",
            icon="save",
            description="Writes text or JSON to a file.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("path", "in", "data", "path", required=False),
                PortSpec("text", "in", "data", "str", required=False),
                PortSpec("data", "in", "data", "any", required=False),
                PortSpec("written_path", "out", "data", "path", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
            ),
            properties=(
                PropertySpec("path", "path", "output.txt", "Output Path"),
                PropertySpec("as_json", "bool", False, "Serialize As JSON"),
            ),
        )

    def execute(self, ctx: ExecutionContext) -> NodeResult:
        path = _pick_path(ctx, input_key="path", property_key="path", node_name="File Write")
        if path.exists() and path.is_dir():
            raise ValueError(f"File Write path must be a file, not a directory: {path}")

        path.parent.mkdir(parents=True, exist_ok=True)
        as_json = bool(ctx.properties.get("as_json", False))
        if as_json:
            payload = ctx.inputs["data"] if "data" in ctx.inputs else ctx.inputs.get("text", "")
            try:
                serialized = json.dumps(payload, indent=2, sort_keys=True, ensure_ascii=True)
            except TypeError as exc:
                raise ValueError(f"File Write could not serialize payload as JSON: {exc}") from exc
            path.write_text(serialized, encoding="utf-8")
        else:
            payload = ctx.inputs.get("text", ctx.inputs.get("data", ""))
            path.write_text("" if payload is None else str(payload), encoding="utf-8")
        return NodeResult(outputs={"written_path": str(path), "exec_out": True})


class EmailSendNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.email_send",
            display_name="Email Send",
            category="Input / Output",
            icon="mail",
            description="Sends a plaintext email using SMTP.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("subject", "in", "data", "str", required=False),
                PortSpec("body", "in", "data", "str", required=False),
                PortSpec("sent", "out", "data", "bool", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
            ),
            properties=(
                PropertySpec("smtp_host", "str", "localhost", "SMTP Host"),
                PropertySpec("smtp_port", "int", 25, "SMTP Port"),
                PropertySpec("username", "str", "", "Username"),
                PropertySpec("password", "str", "", "Password"),
                PropertySpec("sender", "str", "", "Sender"),
                PropertySpec("to", "str", "", "To"),
                PropertySpec("subject", "str", "EA Node Editor Notification", "Subject"),
                PropertySpec("body", "str", "Workflow run completed.", "Body"),
                PropertySpec("use_tls", "bool", False, "Use TLS"),
            ),
        )

    def execute(self, ctx: ExecutionContext) -> NodeResult:
        smtp_host = str(ctx.properties.get("smtp_host", "localhost")).strip()
        smtp_port = int(ctx.properties.get("smtp_port", 25))
        username = str(ctx.properties.get("username", ""))
        password = str(ctx.properties.get("password", ""))
        sender = str(ctx.properties.get("sender", "")).strip()
        recipient = str(ctx.properties.get("to", ""))
        subject = str(ctx.inputs.get("subject", ctx.properties.get("subject", "")))
        body = str(ctx.inputs.get("body", ctx.properties.get("body", "")))
        recipients = _split_recipients(recipient)
        if not smtp_host:
            raise ValueError("Email Send requires SMTP host.")
        if smtp_port <= 0:
            raise ValueError(f"Email Send SMTP port must be a positive integer. Received: {smtp_port}")
        if not sender:
            raise ValueError("Email Send requires sender email address.")
        if not recipients:
            raise ValueError("Email Send requires at least one recipient in 'to'.")
        if username and not password:
            raise ValueError("Email Send requires password when username is provided.")

        message = EmailMessage()
        message["From"] = sender
        message["To"] = ", ".join(recipients)
        message["Subject"] = subject
        message.set_content(body)

        try:
            with smtplib.SMTP(host=smtp_host, port=smtp_port, timeout=10) as smtp:
                if bool(ctx.properties.get("use_tls", False)):
                    smtp.starttls()
                if username:
                    smtp.login(username, password)
                smtp.send_message(message)
        except smtplib.SMTPException as exc:
            raise RuntimeError(f"Email Send SMTP error ({smtp_host}:{smtp_port}): {exc}") from exc
        except OSError as exc:
            raise RuntimeError(
                f"Email Send could not connect to SMTP server {smtp_host}:{smtp_port}: {exc}"
            ) from exc
        return NodeResult(outputs={"sent": True, "exec_out": True})


class ProcessRunNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.process_run",
            display_name="Process Run",
            category="Input / Output",
            icon="terminal",
            description="Executes an external command and captures stdout/stderr.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("command", "in", "data", "str", required=False),
                PortSpec("args", "in", "data", "list[str]", required=False),
                PortSpec("stdin_text", "in", "data", "str", required=False),
                PortSpec("stdout", "out", "data", "str", exposed=True),
                PortSpec("stderr", "out", "data", "str", exposed=True),
                PortSpec("exit_code", "out", "data", "int", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
                PortSpec("on_failed", "out", "failed", "any", exposed=True),
            ),
            properties=(
                PropertySpec("command", "str", "", "Command"),
                PropertySpec("args", "str", "[]", "Args"),
                PropertySpec("cwd", "path", "", "Working Directory"),
                PropertySpec("env", "json", {}, "Environment"),
                PropertySpec("timeout_sec", "float", 60.0, "Timeout (sec)"),
                PropertySpec("shell", "bool", False, "Use Shell"),
                PropertySpec("fail_on_nonzero", "bool", True, "Fail on Non-zero"),
                PropertySpec("encoding", "str", "utf-8", "Encoding"),
            ),
        )

    def execute(self, ctx: ExecutionContext) -> NodeResult:
        command = str(ctx.inputs.get("command", ctx.properties.get("command", ""))).strip()
        if not command:
            raise ValueError("Process Run requires a command.")

        args = _normalize_args(ctx.inputs.get("args", ctx.properties.get("args", "[]")))
        stdin_text = str(ctx.inputs.get("stdin_text", ""))
        cwd_text = str(ctx.properties.get("cwd", "")).strip()
        cwd = cwd_text or None
        if cwd and not Path(cwd).exists():
            raise ValueError(f"Process Run working directory does not exist: {cwd}")

        timeout_sec = float(ctx.properties.get("timeout_sec", 60.0))
        if timeout_sec <= 0:
            raise ValueError("Process Run timeout_sec must be > 0.")

        shell = bool(ctx.properties.get("shell", False))
        fail_on_nonzero = bool(ctx.properties.get("fail_on_nonzero", True))
        encoding = str(ctx.properties.get("encoding", "utf-8")).strip() or "utf-8"
        env_overrides = _normalize_env(ctx.properties.get("env", {}))
        env = dict(os.environ)
        env.update(env_overrides)

        popen_args: Any
        if shell:
            popen_args = " ".join([command, *args]).strip()
        else:
            popen_args = [command, *args]

        process = subprocess.Popen(
            popen_args,
            cwd=cwd,
            env=env,
            shell=shell,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding=encoding,
            errors="replace",
        )

        def _cancel() -> None:
            if process.poll() is not None:
                return
            process.terminate()
            try:
                process.wait(timeout=0.6)
            except subprocess.TimeoutExpired:
                process.kill()
            finally:
                for stream in (process.stdin, process.stdout, process.stderr):
                    try:
                        if stream:
                            stream.close()
                    except OSError:
                        continue

        ctx.register_cancel(_cancel)

        if process.stdin:
            try:
                if stdin_text:
                    process.stdin.write(stdin_text)
                process.stdin.close()
            except OSError:
                pass

        def _append_bounded(
            chunks: deque[str],
            total_chars: int,
            text: str,
        ) -> tuple[int, bool]:
            if not text:
                return total_chars, False
            chunks.append(text)
            total_chars += len(text)
            truncated = False
            while total_chars > _PROCESS_STREAM_CAPTURE_CHAR_LIMIT and chunks:
                total_chars -= len(chunks.popleft())
                truncated = True
            return total_chars, truncated

        stream_queue: queue.Queue[tuple[str, str]] = queue.Queue(maxsize=_PROCESS_STREAM_QUEUE_SIZE)
        dropped_chunks: dict[str, int] = {"stdout": 0, "stderr": 0}
        dropped_lock = threading.Lock()

        def _stream_reader(stream_name: str, stream: Any) -> None:
            try:
                for chunk in iter(stream.readline, ""):
                    if not chunk:
                        break
                    try:
                        stream_queue.put_nowait((stream_name, chunk))
                    except queue.Full:
                        with dropped_lock:
                            dropped_chunks[stream_name] = dropped_chunks.get(stream_name, 0) + 1
            except OSError:
                return

        reader_threads: list[threading.Thread] = []
        if process.stdout is not None:
            stdout_reader = threading.Thread(
                target=_stream_reader,
                args=("stdout", process.stdout),
                daemon=True,
                name="process-run-stdout-reader",
            )
            stdout_reader.start()
            reader_threads.append(stdout_reader)
        if process.stderr is not None:
            stderr_reader = threading.Thread(
                target=_stream_reader,
                args=("stderr", process.stderr),
                daemon=True,
                name="process-run-stderr-reader",
            )
            stderr_reader.start()
            reader_threads.append(stderr_reader)

        started_at = time.monotonic()
        stdout_chunks: deque[str] = deque()
        stderr_chunks: deque[str] = deque()
        stdout_chars = 0
        stderr_chars = 0
        stdout_truncated = False
        stderr_truncated = False
        while True:
            if ctx.should_stop():
                _cancel()
                raise InterruptedError("run_stop_requested")
            elapsed = time.monotonic() - started_at
            if elapsed > timeout_sec:
                _cancel()
                raise TimeoutError(f"Process Run timed out after {timeout_sec:.2f} seconds.")

            drained_any = False
            try:
                while True:
                    stream_name, chunk = stream_queue.get_nowait()
                    drained_any = True
                    message = chunk.rstrip("\r\n")
                    for line in message.splitlines():
                        if line:
                            ctx.emit_log("info", f"[{stream_name}] {line}")
                    if stream_name == "stdout":
                        stdout_chars, was_truncated = _append_bounded(stdout_chunks, stdout_chars, chunk)
                        stdout_truncated = stdout_truncated or was_truncated
                    elif stream_name == "stderr":
                        stderr_chars, was_truncated = _append_bounded(stderr_chunks, stderr_chars, chunk)
                        stderr_truncated = stderr_truncated or was_truncated
            except queue.Empty:
                pass

            if process.poll() is not None and all(not thread.is_alive() for thread in reader_threads):
                if stream_queue.empty():
                    break
                continue
            if not drained_any:
                time.sleep(0.02)

        for thread in reader_threads:
            thread.join(timeout=0.1)

        try:
            while True:
                stream_name, chunk = stream_queue.get_nowait()
                message = chunk.rstrip("\r\n")
                for line in message.splitlines():
                    if line:
                        ctx.emit_log("info", f"[{stream_name}] {line}")
                if stream_name == "stdout":
                    stdout_chars, was_truncated = _append_bounded(stdout_chunks, stdout_chars, chunk)
                    stdout_truncated = stdout_truncated or was_truncated
                elif stream_name == "stderr":
                    stderr_chars, was_truncated = _append_bounded(stderr_chunks, stderr_chars, chunk)
                    stderr_truncated = stderr_truncated or was_truncated
        except queue.Empty:
            pass

        exit_code = int(process.returncode or 0)
        stdout_text = "".join(stdout_chunks)
        stderr_text = "".join(stderr_chunks)

        if stdout_truncated:
            ctx.emit_log(
                "warning",
                (
                    "Process Run stdout capture exceeded limit and was truncated to the most recent "
                    f"{_PROCESS_STREAM_CAPTURE_CHAR_LIMIT} characters."
                ),
            )
        if stderr_truncated:
            ctx.emit_log(
                "warning",
                (
                    "Process Run stderr capture exceeded limit and was truncated to the most recent "
                    f"{_PROCESS_STREAM_CAPTURE_CHAR_LIMIT} characters."
                ),
            )
        with dropped_lock:
            dropped_stdout = int(dropped_chunks.get("stdout", 0))
            dropped_stderr = int(dropped_chunks.get("stderr", 0))
        if dropped_stdout > 0 or dropped_stderr > 0:
            ctx.emit_log(
                "warning",
                (
                    "Process Run dropped streamed chunks due to output backpressure "
                    f"(stdout={dropped_stdout}, stderr={dropped_stderr})."
                ),
            )

        for stream in (process.stdin, process.stdout, process.stderr):
            try:
                if stream:
                    stream.close()
            except OSError:
                continue
        if fail_on_nonzero and exit_code != 0:
            raise RuntimeError(
                f"Process Run returned non-zero exit code {exit_code}. stderr={stderr_text.strip()}"
            )
        return NodeResult(
            outputs={
                "stdout": stdout_text,
                "stderr": stderr_text,
                "exit_code": exit_code,
                "exec_out": True,
            }
        )
