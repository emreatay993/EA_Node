from __future__ import annotations

import csv
import sys
from typing import Any

from ea_node_editor.nodes.builtins.integrations_common import pick_path, require_existing_file
from ea_node_editor.nodes.types import NodeResult, NodeTypeSpec, PortSpec, PropertySpec

try:
    import openpyxl  # type: ignore
except Exception:  # noqa: BLE001
    openpyxl = None


def require_openpyxl(*, node_name: str) -> None:
    if openpyxl is None:
        runtime_mode = "packaged" if bool(getattr(sys, "frozen", False)) else "source"
        install_guidance = (
            "Rebuild package with openpyxl installed in the build environment."
            if runtime_mode == "packaged"
            else "Install with: pip install openpyxl"
        )
        raise RuntimeError(
            f"{node_name} requires optional dependency 'openpyxl' for XLSX support. "
            f"Runtime mode: {runtime_mode}. CSV remains supported without this dependency. "
            f"{install_guidance}"
        )


def normalize_headers(values: tuple[Any, ...] | list[Any]) -> list[str]:
    counters: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values):
        base = str(value).strip() if value is not None else ""
        if not base:
            base = f"column_{index + 1}"
        count = counters.get(base, 0) + 1
        counters[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def normalize_rows_input(rows_input: Any) -> list[dict[str, Any]]:
    if not isinstance(rows_input, list) or not rows_input:
        raise ValueError("Excel Write requires 'rows' as a non-empty list of dictionaries.")
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows_input, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"Excel Write row {index} must be a dictionary.")
        rows.append({str(key): value for key, value in row.items()})
    return rows


def stable_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers = sorted({key for row in rows for key in row})
    if not headers:
        raise ValueError("Excel Write rows must contain at least one column key.")
    return headers


class ExcelReadNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.excel_read",
            display_name="Excel Read",
            category="Input / Output",
            icon="table_view",
            description="Loads rows from Excel or CSV.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("path", "in", "data", "path", required=False),
                PortSpec("rows", "out", "data", "list[dict]", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
            ),
            properties=(
                PropertySpec("path", "path", "", "File Path"),
                PropertySpec("sheet_name", "str", "", "Sheet Name"),
            ),
        )

    def execute(self, ctx) -> NodeResult:  # noqa: ANN001
        path = pick_path(ctx, input_key="path", property_key="path", node_name="Excel Read")
        require_existing_file(path, node_name="Excel Read")
        suffix = path.suffix.lower()

        rows: list[dict[str, Any]] = []
        if suffix == ".csv":
            with path.open("r", encoding="utf-8", newline="") as handle:
                reader = csv.reader(handle)
                header_row = next(reader, None)
                if header_row is not None:
                    headers = normalize_headers(header_row)
                    for source_row in reader:
                        rows.append(
                            {
                                headers[index]: source_row[index] if index < len(source_row) else ""
                                for index in range(len(headers))
                            }
                        )
        elif suffix in {".xlsx", ".xlsm"}:
            require_openpyxl(node_name="Excel Read")
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
            sheet_name = str(ctx.properties.get("sheet_name", "")).strip()
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    available = ", ".join(workbook.sheetnames)
                    raise ValueError(
                        f"Excel Read sheet '{sheet_name}' was not found. Available sheets: {available}"
                    )
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            try:
                iterator = sheet.iter_rows(values_only=True)
                header_row = next(iterator, None)
                if header_row is not None:
                    headers = normalize_headers(list(header_row))
                    for source_row in iterator:
                        values = list(source_row)
                        rows.append(
                            {
                                headers[index]: values[index] if index < len(values) else None
                                for index in range(len(headers))
                            }
                        )
            finally:
                workbook.close()
        else:
            raise ValueError(
                "Excel Read supports only .csv, .xlsx, and .xlsm files. "
                f"Received: {path.suffix or '<no extension>'}"
            )
        return NodeResult(outputs={"rows": rows, "exec_out": True})


class ExcelWriteNodePlugin:
    def spec(self) -> NodeTypeSpec:
        return NodeTypeSpec(
            type_id="io.excel_write",
            display_name="Excel Write",
            category="Input / Output",
            icon="download",
            description="Writes list[dict] rows to CSV/XLSX.",
            ports=(
                PortSpec("exec_in", "in", "exec", "exec", required=False),
                PortSpec("rows", "in", "data", "list[dict]", required=True),
                PortSpec("path", "in", "data", "path", required=False),
                PortSpec("written_path", "out", "data", "path", exposed=True),
                PortSpec("exec_out", "out", "exec", "exec", exposed=True),
            ),
            properties=(
                PropertySpec("path", "path", "output.csv", "Output Path"),
            ),
        )

    def execute(self, ctx) -> NodeResult:  # noqa: ANN001
        rows = normalize_rows_input(ctx.inputs.get("rows"))
        headers = stable_headers(rows)
        path = pick_path(ctx, input_key="path", property_key="path", node_name="Excel Write")
        suffix = path.suffix.lower()
        if suffix not in {".csv", ".xlsx", ".xlsm"}:
            raise ValueError(
                "Excel Write supports only .csv, .xlsx, and .xlsm output formats. "
                f"Received: {path.suffix or '<no extension>'}"
            )
        if path.exists() and path.is_dir():
            raise ValueError(f"Excel Write path must be a file, not a directory: {path}")

        path.parent.mkdir(parents=True, exist_ok=True)
        if suffix == ".csv":
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                writer.writeheader()
                for row in rows:
                    writer.writerow({header: row.get(header) for header in headers})
        else:
            require_openpyxl(node_name="Excel Write")
            workbook = openpyxl.Workbook()  # type: ignore[union-attr]
            try:
                sheet = workbook.active
                sheet.append(headers)
                for row in rows:
                    sheet.append([row.get(header) for header in headers])
                workbook.save(path)
            finally:
                workbook.close()
        return NodeResult(outputs={"written_path": str(path), "exec_out": True})
